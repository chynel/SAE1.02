from openpyxl import load_workbook

# Dictionnaire des mots-clés et de leurs identifiants
mots_cles = {
    "SyntaxError": "1",
    "TypeError": "2",
    "NameError": "3",
    "IndentationError": "4"
}

# Fichier Excel à ouvrir
fichier_excel = "Erreurs_Separees.xlsx"

# Demander à l'utilisateur d'entrer un mot-clé
print("Entrez un mot-clé parmi : SyntaxError, TypeError, NameError, IndentationError")
mot_cle = input("Mot-clé : ").strip()

# Vérifier si le mot-clé est valide
if mot_cle not in mots_cles:
    print("Mot-clé invalide. Veuillez entrer un mot-clé valide.")
else:
    # Obtenir l'identifiant correspondant
    identifiant = mots_cles[mot_cle]
    
    # Charger le fichier Excel
    workbook = load_workbook(fichier_excel)
    sheet = workbook.active  # Supposons que les données sont dans la première feuille
    
    # Collecter toutes les réponses correspondant à l'identifiant
    reponses = set()  # Utiliser un set pour garantir des réponses uniques
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignorer les en-têtes
        id_colonne, reponse = row[5], row[4]  # Identifiant est la 6ᵉ colonne, Réponse est la 5ᵉ colonne
        if str(id_colonne) == identifiant:  # Comparer en tant que chaîne pour éviter les erreurs de type
            reponses.add(reponse)  # Ajouter la réponse dans le set (automatiquement unique)
    
    # Vérifier si des réponses ont été trouvées
    if reponses:
        # Convertir le set en liste pour pouvoir afficher
        reponses = list(reponses)
        print(f"Voici quelques réponses pour le mot-clé '{mot_cle}':")
        for i, reponse in enumerate(reponses[:20], start=1):  # Limite à 20 réponses
            print(f"{i}. {reponse}")
    else:
        print(f"Aucune réponse trouvée pour le mot-clé '{mot_cle}'.")
