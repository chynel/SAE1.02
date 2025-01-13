from openpyxl import load_workbook, Workbook

# Chemins relatifs pour les fichiers
input_file = "Infos_Etudiants.xlsx"  # Le fichier source doit être dans le même répertoire que le script
output_file = "Erreurs_Separees.xlsx"  # Le fichier de sortie sera généré dans le même répertoire

# Charger le classeur Excel
workbook_source = load_workbook(input_file)
feuille_source = workbook_source.active

# Créer un nouveau classeur Excel pour le résultat
workbook_result = Workbook()
feuille_resultat = workbook_result.active
feuille_resultat.title = "Erreurs"

# En-têtes pour le fichier de sortie (ajout de id_type_erreur)
entetes = ['DATE', 'NUMERO', 'TYPE', 'SPECIFICITE', 'REPONSE', 'id_type_erreur']
feuille_resultat.append(entetes)

# Fonction pour nettoyer les champs
def nettoyer_champ(champ):
    if champ.startswith("="):  # Si le champ ressemble à une formule
        return "'" + champ  # Ajouter une apostrophe pour forcer Excel à le traiter comme texte
    return champ[:1000]  # Tronquer si la longueur dépasse 32 767 caractères

# Fonction pour supprimer les espaces
def supprime_espace(chaine):
    debut = 0
    fin = len(chaine) - 1
    
    while debut <= fin and chaine[debut] == " ":
        debut += 1
    
    while fin >= debut and chaine[fin] == " ":
        fin -= 1
    
    if debut <= fin:
        return chaine[debut:fin+1]
    else:
        return ""

# Fonction pour séparer les champs
def extraire_champs(ligne):
    champs = []
    champ = ""
    for char in ligne:
        if char == "-" and champ:  # Si un champ se termine par " - " et champ n'est pas vide
            champs.append(supprime_espace(champ))  # Ajouter le champ après avoir supprimé les espaces inutiles
            champ = ""  # Réinitialiser le champ
        else:
            champ += char  # Ajouter le caractère au champ
    if champ:  # Ajouter le dernier champ si non vide
        champs.append(supprime_espace(champ))
    return champs

# Fonction pour déterminer l'id_type_erreur
def obtenir_id_type_erreur(type_erreur):
    correspondance = {
        "SyntaxError": "1",
        "TypeError": "2",
        "NameError": "3",
        "IndentationError": "4",
    }
    return correspondance.get(type_erreur, "")  # Retourne une chaîne vide si non trouvé

# Compteur pour suivre les lignes traitées et ignorées
lignes_traitees = 0
lignes_ignorees = 0

# Parcourir les lignes du fichier source
for row in feuille_source.iter_rows(min_row=2, values_only=True):  # Ignorer les en-têtes
    for partie in row:
        if partie:  # Vérifier si la cellule contient une valeur
            # Remplacer ':' par ' - '
            partie = partie.replace(":", " - ")
            
            if " - " in partie:
                champs = extraire_champs(partie)
                if len(champs) >= 5:
                    # Ajouter les champs avec id_type_erreur
                    type_erreur = champs[2]  # TYPE est dans la troisième position
                    id_type_erreur = obtenir_id_type_erreur(type_erreur)
                    ligne_resultat = [nettoyer_champ(champ) for champ in champs[:5]] + [id_type_erreur]
                    feuille_resultat.append(ligne_resultat)
                    lignes_traitees += 1
                else:
                    lignes_ignorees += 1  # Lignes ignorées si elles ont moins de 5 champs
                    print(f"Ligne ignorée (moins de 5 champs) : {partie}")
            else:
                lignes_ignorees += 1  # Lignes ignorées si elles ne contiennent pas " - "
                print(f"Ligne ignorée (pas de '-') : {partie}")
    
# Sauvegarder le fichier Excel résultant après ajout des données
workbook_result.save(output_file)

# Afficher le nombre de lignes traitées et ignorées
print(f"Nombre de lignes traitées : {lignes_traitees}")
print(f"Nombre de lignes ignorées : {lignes_ignorees}")
print(f"Les données ont été extraites et sauvegardées dans le fichier {output_file}.")
